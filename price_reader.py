from openpyxl import load_workbook


def read_price_data(excel_path):
    workbook = load_workbook(excel_path, data_only=True)

    sheet = workbook["All_Items"]

    items = []

    for row in range(2, sheet.max_row + 1):
        name = sheet[f"A{row}"].value
        category = sheet[f"B{row}"].value
        unit = sheet[f"C{row}"].value
        price = sheet[f"D{row}"].value

        if not name:
            continue

        items.append({
            "name": str(name).strip(),
            "category": str(category).strip() if category else "REVIEW",
            "unit": str(unit).strip() if unit else "REVIEW",
            "price": float(price) if price not in (None, "") else 0.0,
        })

    categories_sheet = workbook["Categories"]

    markups = {}

    for row in range(2, categories_sheet.max_row + 1):
        category = categories_sheet[f"A{row}"].value
        coefficient = categories_sheet[f"B{row}"].value

        if not category or coefficient is None:
            continue

        category = str(category)

        if "(" in category and ")" in category:
            category = category.split("(")[-1].split(")")[0]

        markups[category.strip().lower()] = float(coefficient)

    settings = workbook["Global_Settings"]

    project_multiplier = float(settings["B2"].value or 1.0)

    return {
        "items": items,
        "markups": markups,
        "project_multiplier": project_multiplier,
    }